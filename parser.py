import pdfplumber
import re
import spacy
import os
from openpyxl import Workbook, load_workbook
import PyPDF2

nlp = spacy.load("en_core_web_sm")

# Skills list: isme aur add bhi kar sakta hai
skills_list = [
    "Python", "Java", "C++", "C", "HTML", "CSS", "JavaScript", "React",
    "Node", "Flask", "Django", "SQL", "MongoDB", "MySQL", "Git", "Linux",
    "Machine Learning", "Deep Learning", "Data Science", "Pandas", "NumPy",
    "TensorFlow", "Keras", "OpenCV", "AWS", "Docker", "Kubernetes"
]

# PDF se text extract
def extract_text_from_pdf(file_path):
    text = ""
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text

def extract_name(text):
    import re

    # Get top 3 non-empty lines
    lines = text.strip().split('\n')
    lines = [line.strip() for line in lines if line.strip()]
    if not lines:
        return "Not Found"

    first_line = lines[0]

    # Remove phone numbers, emails, numbers, symbols
    first_line = re.sub(r'[\+\-\d\@\:\|•©§&%$!#\^*\(\)\[\]{}<>=_,]', '', first_line)
    first_line = re.sub(r'\s{2,}', ' ', first_line).strip()

    # Extract capitalized words from cleaned first line
    words = [word for word in first_line.split() if word[0].isupper()]

    # Return first 2–3 words only as name
    if 1 < len(words) <= 4:
        return ' '.join(words[:3])  # name likely to be max 3 parts
    return "Not Found"


# Email extract
def extract_email(text):
    match = re.search(r'\S+@\S+', text)
    return match.group() if match else "Not Found"

# Phone extract
def extract_phone(text):
    match = re.search(r'(\+?\d[\d\s\-]{8,15})', text)
    return match.group() if match else "Not Found"

# Resume se skills extract
def extract_skills(text):
    text_lower = text.lower()
    found = []
    for skill in skills_list:
        if re.search(r'\b' + re.escape(skill.lower()) + r'\b', text_lower):
            found.append(skill)
    return found

# Matching logic
def calculate_match_percentage(resume_skills, job_desc):
    if not job_desc or not resume_skills:
        return 0

    job_desc = job_desc.lower()
    match_count = 0

    for skill in resume_skills:
        if re.search(r'\b' + re.escape(skill.lower()) + r'\b', job_desc):
            match_count += 1

    return round((match_count / len(resume_skills)) * 100, 2) if resume_skills else 0

# def extract_links(text):
#     # Match all http/https links
#     links = re.findall(r'https?://[^\s\)\]\}\'\"<]+', text)
#     return links

def extract_links(pdf_path):
        links = []
        with open(pdf_path, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                if '/Annots' in page:
                    for annot in page['/Annots']:
                        annot_obj = annot.get_object()
                        if '/A' in annot_obj and '/URI' in annot_obj['/A']:
                            url = annot_obj['/A']['/URI']
                            links.append(url)
        return links

def save_to_excel(data, file_path="data/parsed_data.xlsx"):
    headers = ["Name", "Email", "Phone", "Skills", "Links", "Match %"]

    # Create new workbook if file doesn't exist
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    else:
        wb = load_workbook(file_path)
        ws = wb.active

    # Append new row
    links = ', '.join(data.get("links", []))
    skills = ', '.join(data.get("skills", []))
    ws.append([
        data.get("name", ""),
        data.get("email", ""),
        data.get("phone", ""),
        skills,
        links,
        data.get("match_percentage", "") if "match_percentage" in data else ""
    ])

    wb.save(file_path)

# Final combined function
def parse_resume(file_path, job_desc=None):
    text = extract_text_from_pdf(file_path)
 

    skills = extract_skills(text)
    name = extract_name(text)
    email = extract_email(text)
    phone = extract_phone(text)
    links = extract_links(file_path)

    result = {
        "name": name,
        "email": email,
        "phone": phone,
        "skills": skills,
        "links": links
    }


    if job_desc:
        result["match_percentage"] = calculate_match_percentage(skills, job_desc)

    # Save to Excel
    save_to_excel(result)

    return result   



